import pandas as pd
from timing import timeit
from openpyxl import Workbook
from openpyxl.styles import Font

@timeit
def get_excel(excursion_destinations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Destinations"

    ws.append(["Name", "Description", "Contact"])

    counter = 2
    for dest in excursion_destinations:
        ws[f'A{counter}'].hyperlink = dest.url
        ws[f'A{counter}'].value = dest.name
        ws[f'A{counter}'].style = "Hyperlink"
        ws[f'B{counter}'].value = dest.description
        ws[f'C{counter}'].value = str.join(", ", dest.contact_info)
        counter += 1

    ft = Font(bold=True)
    for row in ws["A1:C1"]:
        for cell in row:
            cell.font = ft

    wb.save('./output/destinations.xlsx')